from copy import copy
from langdetect import detect, DetectorFactory
from openpyxl import load_workbook
from pybtex.database.input import bibtex
from pybtex.database.output.bibtex import Writer
from pybtex.database import parse_file
from transliterate import translit

DetectorFactory.seed = 0

# The default for this library is to use quotation marks instead of curly braces
class CustomBibtexWriter(Writer):
    def quote(self, s):
        return '{' + s + '}'

# Method for styling hyperlink cells when adding them
def hyperlinkStyling(current_row):
    # set the last cell to be a hyperlink
    cell = sheet.cell(row=current_row, column=5)
    left_cell = sheet.cell(row=current_row, column=4)

    # styling for hyperlink
    new_font = copy(left_cell.font)
    new_font.underline = "single"
    new_font.color = "0000FF"  # blue color in hex
    cell.font = new_font
    cell.hyperlink = link

bibtex_file = 'test.bib'
excel_file = "test.xlsx"

parser = bibtex.Parser()
bib_data = parser.parse_file(bibtex_file)

bibtex_dict = {}
for entry_key, entry in bib_data.entries.items():
    raw_orcids = entry.fields['orcid'].split(" and ")
    orcids = [orcid if len(orcid.strip()) >= 19 else None for orcid in raw_orcids] # orcid is 19 char long, > is added in case formatting makes it seem longer
    for i, author in enumerate(entry.persons['author']):
        first = " ".join(author.first_names) if author.first_names else ''
        last = " ".join(author.last_names) if author.last_names else ''
        first = first.replace("Ð", "Đ").replace("ð", "đ") # Ð and Đ are of different unicode values...
                                                                                   # Poludeo sam tražeći problem haha
        last = last.replace("Ð", "Đ").replace("ð", "đ")
        if len(orcids) > i :
            bibtex_dict[(first, last)] = orcids[i]
        else:
            bibtex_dict[(first, last)] = None

wb = load_workbook(excel_file)
sheet = wb.active

# Remove empty rows from the sheet
for row in range(sheet.max_row, 0, -1):
    if all(cell.value in (None, "") for cell in sheet[row]):
        sheet.delete_rows(row)

headers = [cell.value for cell in sheet[1]]  # Header row
headers = [s.strip() if s is not None else None for s in headers] #strip header whitespace

# indexes of the rows
firstname_index = headers.index("Име")
lastname_index = headers.index("Презиме")
orcid_index = headers.index("ORCId идентификатор")

excel_dict = {}
i = 2
# Iterate over each row
for row in sheet.iter_rows(min_row=2, values_only=True):
    first_cyr = row[firstname_index]  # First name in Cyrillic
    last_cyr = row[lastname_index]  # Last name in Cyrillic
    orcid = row[orcid_index]  # orcid
    if first_cyr is None or last_cyr is None:
        continue
    # turn cyrillic to latin
    first_lat = translit(first_cyr, 'sr', reversed=True)
    last_lat = translit(last_cyr, 'sr', reversed=True)
    first_lat = first_lat.replace("Ð", "Đ").replace("ð", "đ")  # Ð and Đ are of different unicode values...
    last_lat = last_lat.replace("Ð", "Đ").replace("ð", "đ")

    # If the bibtex file had the orcid for this person, but the excel doesn't, add it
    if orcid is None and bibtex_dict.keys().__contains__((first_lat, last_lat)) and bibtex_dict[(first_lat, last_lat)] is not None:
        sheet.cell(row=i, column=orcid_index + 1).value = bibtex_dict[(first_lat, last_lat)]
        link = f"https://orcid.org/{bibtex_dict[(first_lat, last_lat)]}"
        sheet.cell(row=i, column=orcid_index + 2).value = link
        hyperlinkStyling(i)
        print("Modified " + first_lat + " " + last_lat + " from Bibtex to excel")
    excel_dict[(first_lat, last_lat)] = orcid
    i += 1

noidbib = set() # Keep track of which names don't have orcids
for name in bibtex_dict.keys():
    if bibtex_dict[name] is None:
        noidbib.add(name)

bibset = set(bibtex_dict.keys())
excelset = set(excel_dict.keys())
common_names = bibset & excelset # List of names in both the bibtex and Excel
modify_bibtex = {}

for name in common_names:
    if bibtex_dict[name] is None and excel_dict[name] is not None: # There is data in the excel, but not bibtex
        modify_bibtex[name] = excel_dict[name]

# add excel data to bibtex
for entry_key, entry in bib_data.entries.items():
    author_list = entry.persons['author'] # get the total number of authors
    will_add = False
    raw_orcids = entry.fields['orcid'].split(" and ")
    orcids = [orcid if len(orcid.strip()) >= 19 else None for orcid in raw_orcids]


    # If a orcid field is empty and we aren't adding data, keep it empty, don't add "nooorcid"
    # If an entry has multiple authors, and we only add info for one, we should see something like
    # "noorcid and {orcid}" instead of just one id
    # this is why we need to make an orcids list that will modify the orcid field after we process all the authors

    for i, author in enumerate(author_list):
        first = " ".join(author.first_names) if author.first_names else ''
        last = " ".join(author.last_names) if author.last_names else ''
        first = first.replace("Ð", "Đ").replace("ð", "đ")  # Ð and Đ are of different unicode values...
        last = last.replace("Ð", "Đ").replace("ð", "đ")
        name = (first, last)
        if modify_bibtex.__contains__(name): # if the author is in the update dict, we update the entry
            if len(orcids) <= i:
                orcids.append(modify_bibtex[name])
            else :
                orcids[i] = modify_bibtex[name]
            print("modified " + name[0] + " " + name[1] + " in the bibtex file")
        elif name in noidbib: # check to see if we need to add "noorc" to the field
            if len(orcids) <= i:
                orcids.append(None)
            else:
                orcids[i] = None

    # only update entry if there are ids to update
    emptyfield = True
    for orcid in orcids:
        if orcid is not None and orcid not in modify_bibtex.keys():
            emptyfield = False

    # If there are ids, update the entry
    if len(orcids) > 0 and not emptyfield:
        orcid_string = ""
        if orcids[0] is None:
            orcid_string = "noorcid"
        else: # fence posting
            orcid_string = orcids[0]
        for orcid in orcids[1:]:
            if orcid is not None:
                orcid_string = orcid_string + " and " + orcid
            else:
                orcid_string = orcid_string + " and " + "noorcid"
        entry.fields['orcid'] = orcid_string



# People in the bibtex, not in the Excel
add_to_excel = set(bibtex_dict.keys()) - set(excel_dict.keys())
for name in add_to_excel:
    # Transliterating into Cyrillic causes issues, here is somewhat of a solution
    if detect(name[0] + " " + name[1]) in ['hr', 'sl']:  # if this classifies names as Croatian or Slovenian, transcribe
        first = translit(name[0], 'sr', reversed=False)
        last = translit(name[1], 'sr', reversed=False)
    else:
        first = name[0]
        last = name[1]
        # print(name[0] + " " + name[1] + " or " + first_cyr + " " + last_cyr + " is " + detect(name[0] + " " + name[1])) # testing

    orcid = bibtex_dict[name]
    link = ""
    if orcid is not None:
        link = f"https://orcid.org/{orcid}"
    sheet.append(["", first, last, orcid, link])
    print("added " + first + " " + last + " to the excel file")
    last_row = sheet.max_row
    hyperlinkStyling(last_row)

# Save excel changes a desired output file
destination_file = "test.xlsx"
wb.save(destination_file)

# Save bibtex changes to a desired output file
writer = CustomBibtexWriter()
modified_bibtex_file = "test.bib"
writer.write_file(bib_data, modified_bibtex_file)


bib_data = parse_file('test.bib')

# max field-name length
max_field = max(
    len(field)
    for entry in bib_data.entries.values()
    for field in entry.fields
)

# Turns the Person objects into a single string
def format_authors(persons):
    names = []
    for person in persons:
        first = " ".join(person.first_names)
        last  = " ".join(person.last_names)
        names.append(f"{last}, {first}")
    return " and ".join(names)

# To format the output file with the exact same spacing as the input
with open('test.bib', 'w', encoding='utf8') as out:
    for key, entry in bib_data.entries.items():
        out.write(f"@{entry.type}{{{key},\n")
        author_str = format_authors(entry.persons.get('author', []))
        padding = ' ' * (max_field - len('author') + 1)
        out.write(f"  author{padding}= {{{author_str}}},\n")
        for field, value in entry.fields.items():
            pad = ' ' * (max_field - len(field) + 1)
            out.write(f"  {field}{pad} = {{{value}}},\n")
        out.write("}\n\n")

with open('test.bib', 'r', encoding='utf8') as f:
    text = f.read()
# For some reason backslashes are added to the output, seems to be due to the library
text = text.replace(r'\&', '&')
with open('test.bib', 'w', encoding='utf8') as f:
    f.write(text)
